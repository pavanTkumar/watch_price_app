import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font

class WatchPricingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Watch Service Price Manager")
        
        self.setup_styles()
        self.setup_window_size()
        
        # Allow window to be resizable
        self.root.resizable(True, True)
        
        # Configure weight of rows and columns for responsive layout
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        
        self.selected_item = None
        self.get_excel_file_location()
        
        if self.excel_file:
            self.setup_fresh_excel_file()
            self.create_main_interface()
            self.load_existing_data()
        else:
            self.root.destroy()

    def setup_styles(self):
        style = ttk.Style()
        style.configure("Header.TLabel", font=('Arial', 16, 'bold'))
        style.configure("Subheader.TLabel", font=('Arial', 12, 'bold'))
        style.configure("Footer.TLabel", font=('Arial', 9, 'bold'), foreground='blue')

    def setup_window_size(self):
        # Get screen dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Calculate window size (80% of screen size)
        self.window_width = int(screen_width * 0.8)
        self.window_height = int(screen_height * 0.8)
        
        # Calculate position for center of screen
        position_top = int(screen_height/2 - self.window_height/2)
        position_right = int(screen_width/2 - self.window_width/2)
        
        # Set the position of the window
        self.root.geometry(f"{self.window_width}x{self.window_height}+{position_right}+{position_top}")

    def get_excel_file_location(self):
        # First ask if user wants to open existing file
        response = messagebox.askyesno(
            "File Selection",
            "Do you want to open an existing price manager file?\n\n" +
            "Click 'Yes' to open existing file\n" +
            "Click 'No' to create new file"
        )
        
        if response:  # User wants to open existing file
            file_path = filedialog.askopenfilename(
                title="Select Price Manager File",
                filetypes=[("Excel files", "*.xlsx")],
                initialdir=os.path.expanduser("~/Documents")
            )
            if file_path:  # User selected a file
                self.excel_file = file_path
            else:  # User cancelled file selection
                self.excel_file = None
        else:  # User wants to create new file
            file_path = filedialog.asksaveasfilename(
                title="Save Price List As",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialdir=os.path.expanduser("~/Documents"),
                initialfile="Enter File Name.xlsx"
            )
            if file_path:  # User selected a location and filename
                self.excel_file = file_path
            else:  # User cancelled file selection
                self.excel_file = None
    def setup_fresh_excel_file(self):
        # Only create new file if it doesn't exist
        if not os.path.exists(self.excel_file):
            wb = openpyxl.Workbook()
            headers = ["Brand", "Price", "Category", "Service Type", "Date Added"]
            sheet = wb.active
            sheet.title = "Watch_Services"
            sheet.append(headers)
            for col in range(1, len(headers) + 1):
                sheet.cell(row=1, column=col).font = Font(bold=True)
            wb.save(self.excel_file)

    def create_main_interface(self):
        # Main container frame
        main_container = ttk.Frame(self.root)
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Configure weights for responsive layout
        main_container.grid_columnconfigure(0, weight=1)
        main_container.grid_rowconfigure(1, weight=1)

        # Header - Updated text
        header = ttk.Label(main_container, text="Price Manager for Our Time Jewelers", style="Header.TLabel")
        header.pack(pady=(0, 20))

        # Input Frame with responsive layout
        input_frame = ttk.LabelFrame(main_container, text="Add New Entry", padding=10)
        input_frame.pack(fill=tk.X, pady=(0, 10))

        # Service Type Dropdown with Overhaul
        service_frame = ttk.Frame(input_frame)
        service_frame.pack(fill=tk.X, pady=5)
        ttk.Label(service_frame, text="Service Type:").pack(side=tk.LEFT, padx=5)
        self.service_type = ttk.Combobox(
            service_frame,
            values=["5 Year Battery", "Lifetime Battery", "Band Adjustment", "Overhaul"],
            state="readonly",
            width=min(30, self.window_width // 40)
        )
        self.service_type.set("5 Year Battery")
        self.service_type.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # Category Frame with Add Category button
        category_frame = ttk.Frame(input_frame)
        category_frame.pack(fill=tk.X, pady=5)
        
        # Left side - Category dropdown
        cat_left_frame = ttk.Frame(category_frame)
        cat_left_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Label(cat_left_frame, text="Category:").pack(side=tk.LEFT, padx=5)
        self.category_var = tk.StringVar()
        self.category_type = ttk.Combobox(
            cat_left_frame,
            textvariable=self.category_var,
            width=min(30, self.window_width // 40)
        )
        self.service_type.bind('<<ComboboxSelected>>', self.update_categories)
        self.category_type.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # Right side - Add Category button
        ttk.Button(category_frame, text="Add Category", 
                  command=self.add_new_category).pack(side=tk.RIGHT, padx=5)

        # Set initial categories
        self.update_categories(None)

        # Brand Entry
        brand_frame = ttk.Frame(input_frame)
        brand_frame.pack(fill=tk.X, pady=5)
        ttk.Label(brand_frame, text="Watch Brand:").pack(side=tk.LEFT, padx=5)
        self.brand_entry = ttk.Entry(brand_frame)
        self.brand_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # Price Entry
        price_frame = ttk.Frame(input_frame)
        price_frame.pack(fill=tk.X, pady=5)
        ttk.Label(price_frame, text="Price ($):").pack(side=tk.LEFT, padx=5)
        self.price_entry = ttk.Entry(price_frame)
        self.price_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # Buttons Frame
        button_frame = ttk.Frame(input_frame)
        button_frame.pack(pady=10)
        
        # Create buttons with dynamic padding
        pad_x = max(5, min(20, self.window_width // 100))
        ttk.Button(button_frame, text="Add Entry", command=self.add_entry).pack(side=tk.LEFT, padx=pad_x)
        ttk.Button(button_frame, text="Update Selected", command=self.update_entry).pack(side=tk.LEFT, padx=pad_x)
        ttk.Button(button_frame, text="Remove Selected", command=self.remove_entry).pack(side=tk.LEFT, padx=pad_x)

        # Tree Frame (responsive)
        tree_frame = ttk.LabelFrame(main_container, text="Entries", padding=10)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        # Create Treeview with dynamic column widths
        self.tree = ttk.Treeview(
            tree_frame,
            columns=("Brand", "Price", "Category", "Service", "Date"),
            show="headings"
        )
        
        # Calculate relative column widths
        total_width = self.window_width - 100
        self.tree.column("Brand", width=int(total_width * 0.3))
        self.tree.column("Price", width=int(total_width * 0.15))
        self.tree.column("Category", width=int(total_width * 0.15))
        self.tree.column("Service", width=int(total_width * 0.2))
        self.tree.column("Date", width=int(total_width * 0.2))

        # Configure headings
        for col in ("Brand", "Price", "Category", "Service", "Date"):
            self.tree.heading(col, text=col)

        # Bind select event
        self.tree.bind('<<TreeviewSelect>>', self.on_select)

        # Scrollbars
        y_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        y_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        x_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        x_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.tree.configure(yscrollcommand=y_scrollbar.set, xscrollcommand=x_scrollbar.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        # Add Footer at the bottom
        footer_frame = ttk.Frame(main_container)
        footer_frame.pack(side=tk.BOTTOM, pady=(10, 0))
        
        footer_label = ttk.Label(
            footer_frame, 
            text="Developed By Pavan Tejavath\nwww.thetejavath.com", 
            style="Footer.TLabel",
            cursor="hand2"  # Changes cursor to hand when hovering
        )
        footer_label.pack()
        
        # Make the footer clickable
    def open_website(event):
        import webbrowser
        webbrowser.open('http://www.thetejavath.com')
        footer_label.bind('<Button-1>', open_website)
        # Bind resize event
        self.root.bind('<Configure>', self.on_window_resize)
    def add_new_category(self):
        # Create a new window for category input
        dialog = tk.Toplevel(self.root)
        dialog.title("Add New Category")
        dialog.geometry("300x150")
        
        # Center the dialog
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Add input field
        ttk.Label(dialog, text="Enter new category name:").pack(pady=10)
        entry = ttk.Entry(dialog, width=40)
        entry.pack(pady=5, padx=10)
        entry.focus()

        def save_category():
            new_category = entry.get().strip()
            if new_category:
                current_categories = list(self.category_type['values'])
                if not current_categories or new_category not in current_categories:
                    new_categories = list(current_categories) if current_categories else []
                    new_categories.append(new_category)
                    self.category_type['values'] = new_categories
                    self.category_type.set(new_category)
                    dialog.destroy()
                    messagebox.showinfo("Success", f"Added new category: {new_category}")
                else:
                    messagebox.showwarning("Warning", "Category already exists!")
            else:
                messagebox.showwarning("Warning", "Please enter a category name")

        # Add buttons
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=20)
        ttk.Button(button_frame, text="Save", command=save_category).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=5)

    def update_categories(self, event):
        service = self.service_type.get()
        current_categories = list(self.category_type['values']) if self.category_type['values'] else []
        
        if service == "5 Year Battery":
            default_categories = ["Category 1", "Category 2", "Category 3", "Category 4", "Category 5"]
        elif service == "Lifetime Battery":
            default_categories = ["Category 1", "Category 2", "Category 3", "Category 4", "Category 5"]
        elif service == "Band Adjustment":
            default_categories = ["Basic", "Mid-Range", "High-End"]
        else:  # Overhaul
            default_categories = ["Basic Service", "Full Service", "Complete Restoration"]

        # Combine default categories with any custom categories
        all_categories = list(set(default_categories + current_categories))
        all_categories.sort()
        
        self.category_type['values'] = all_categories
        if all_categories:
            self.category_type.set(all_categories[0])

    def check_duplicate(self, brand, service_type, current_date=None):
        # Get all items from treeview
        for item in self.tree.get_children():
            values = self.tree.item(item)['values']
            # If updating, ignore the current entry being updated
            if current_date and values[4] == current_date:
                continue
            # Check if brand and service type match
            if values[0].lower() == brand.lower() and values[3] == service_type:
                return True
        return False

    def add_entry(self):
        try:
            brand = self.brand_entry.get().strip()
            price = self.price_entry.get().strip()
            service_type = self.service_type.get()
            category = self.category_type.get()

            if not brand or not price or not category:
                messagebox.showerror("Error", "Please fill in all fields")
                return

            try:
                price_float = float(price)
                if price_float <= 0:
                    messagebox.showerror("Error", "Price must be greater than 0")
                    return
            except ValueError:
                messagebox.showerror("Error", "Invalid price format")
                return

            # Check for duplicates
            if self.check_duplicate(brand, service_type):
                response = messagebox.askyesno(
                    "Duplicate Entry",
                    f"A record for {brand} with {service_type} already exists.\nDo you want to add it anyway?"
                )
                if not response:
                    return

            date_added = datetime.now().strftime("%Y-%m-%d %H:%M")

            # Add to treeview
            self.tree.insert("", 0, values=(brand, f"${price}", category, service_type, date_added))

            # Add to Excel
            wb = openpyxl.load_workbook(self.excel_file)
            sheet = wb.active
            sheet.append([brand, float(price), category, service_type, date_added])
            wb.save(self.excel_file)

            # Clear entries
            self.brand_entry.delete(0, tk.END)
            self.price_entry.delete(0, tk.END)
            self.brand_entry.focus()

            messagebox.showinfo("Success", f"Added {brand} to {category} ({service_type})")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def on_select(self, event):
        selected_items = self.tree.selection()
        if selected_items:
            self.selected_item = selected_items[0]
            values = self.tree.item(self.selected_item)['values']
            self.brand_entry.delete(0, tk.END)
            self.brand_entry.insert(0, values[0])
            self.price_entry.delete(0, tk.END)
            self.price_entry.insert(0, values[1].replace('$', ''))
            self.service_type.set(values[3])
            # Update categories and set the selected category
            self.update_categories(None)
            self.category_type.set(values[2])
        else:
            self.selected_item = None
    def update_entry(self):
        if not self.selected_item:
            messagebox.showwarning("Warning", "Please select an item to update")
            return

        try:
            brand = self.brand_entry.get().strip()
            price = self.price_entry.get().strip()
            service_type = self.service_type.get()
            category = self.category_type.get()
            current_values = self.tree.item(self.selected_item)['values']

            if not brand or not price or not category:
                messagebox.showerror("Error", "Please fill in all fields")
                return

            try:
                price_float = float(price)
                if price_float <= 0:
                    messagebox.showerror("Error", "Price must be greater than 0")
                    return
            except ValueError:
                messagebox.showerror("Error", "Invalid price format")
                return

            # Check for duplicates if brand or service type changed
            if (brand.lower() != current_values[0].lower() or 
                service_type != current_values[3]):
                if self.check_duplicate(brand, service_type, current_values[4]):
                    response = messagebox.askyesno(
                        "Duplicate Entry",
                        f"A record for {brand} with {service_type} already exists.\nDo you want to update anyway?"
                    )
                    if not response:
                        return

            date_added = current_values[4]

            # Update treeview
            self.tree.item(self.selected_item, values=(brand, f"${price}", category, service_type, date_added))

            # Update Excel file
            wb = openpyxl.load_workbook(self.excel_file)
            sheet = wb.active
            
            for row in sheet.iter_rows(min_row=2):
                if (row[0].value == current_values[0] and 
                    row[4].value == current_values[4]):
                    row[0].value = brand
                    row[1].value = float(price)
                    row[2].value = category
                    row[3].value = service_type
                    break
            
            wb.save(self.excel_file)

            # Clear entries
            self.brand_entry.delete(0, tk.END)
            self.price_entry.delete(0, tk.END)
            self.service_type.set("5 Year Battery")
            self.selected_item = None

            messagebox.showinfo("Success", f"Updated {brand}")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def remove_entry(self):
        if not self.selected_item:
            messagebox.showwarning("Warning", "Please select an item to remove")
            return

        if not messagebox.askyesno("Confirm", "Are you sure you want to remove this entry?"):
            return

        try:
            values = self.tree.item(self.selected_item)['values']
            
            # Remove from treeview
            self.tree.delete(self.selected_item)

            # Remove from Excel file
            wb = openpyxl.load_workbook(self.excel_file)
            sheet = wb.active
            
            # Find and remove the row in Excel
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                if (row[0].value == values[0] and 
                    row[4].value == values[4]):  # Match brand and date to identify unique row
                    sheet.delete_rows(row_idx)
                    break
            
            wb.save(self.excel_file)

            # Clear entries
            self.brand_entry.delete(0, tk.END)
            self.price_entry.delete(0, tk.END)
            self.service_type.set("5 Year Battery")
            self.selected_item = None

            messagebox.showinfo("Success", "Entry removed successfully")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def on_window_resize(self, event):
        # Only handle window resize events, not widget resize events
        if event.widget == self.root:
            # Update stored dimensions
            self.window_width = event.width
            self.window_height = event.height
            
            # Update tree column widths
            total_width = self.window_width - 100
            self.tree.column("Brand", width=int(total_width * 0.3))
            self.tree.column("Price", width=int(total_width * 0.15))
            self.tree.column("Category", width=int(total_width * 0.15))
            self.tree.column("Service", width=int(total_width * 0.2))
            self.tree.column("Date", width=int(total_width * 0.2))

    def load_existing_data(self):
        if os.path.exists(self.excel_file):
            wb = openpyxl.load_workbook(self.excel_file)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and len(row) >= 5:  # Ensure row has all required fields
                    self.tree.insert("", 0, values=(
                        row[0],  # Brand
                        f"${row[1]}" if row[1] else "",  # Price
                        row[2] if row[2] else "",  # Category
                        row[3] if row[3] else "",  # Service Type
                        row[4] if row[4] else ""   # Date
                    ))

def main():
    root = tk.Tk()
    app = WatchPricingApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()